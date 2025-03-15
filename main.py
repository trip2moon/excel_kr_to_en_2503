import asyncio
import copy
import json
import os

import openpyxl
from googletrans import Translator

DIR_EXCEL_INPUT = 'excel_input'
DIR_EXCEL_OUTPUT = 'excel_output'
DIR_REFERENCE = 'reference'
DIR_EXCEL_TARGET = 'excel_target'

FILE_EXCEL_EXT = 'xlsx'
FILE_EXCEL_NAME = '엑셀_헌법_20250316'
FILE_EXCEL_INPUT = f'{FILE_EXCEL_NAME}.{FILE_EXCEL_EXT}'
FILE_EXCEL_OUTPUT = f'{FILE_EXCEL_NAME}.영문.{FILE_EXCEL_EXT}'

SHEET_NAME_INPUT = '헌법'
SHEET_NAME_OUTPUT = f'{SHEET_NAME_INPUT}.영문'

FILE_REFERENCE_JSON = f'reference_{SHEET_NAME_INPUT}_translated.json'
FILE_REFERENCE_TARGET_JSON = f'reference_target_translated.json'

FONT_NAME_EN = 'Consolas'


def get_file_path_input_output():
    if os.path.exists(DIR_EXCEL_OUTPUT) is False:
        os.makedirs(DIR_EXCEL_OUTPUT)
    file_path_input = os.path.join(DIR_EXCEL_INPUT, FILE_EXCEL_INPUT)
    file_path_output = os.path.join(DIR_EXCEL_OUTPUT, FILE_EXCEL_OUTPUT)
    return file_path_input, file_path_output


def get_file_path_reference():
    if os.path.exists(DIR_REFERENCE) is False:
        os.makedirs(DIR_REFERENCE)
    file_path = os.path.join(DIR_REFERENCE, FILE_REFERENCE_TARGET_JSON)
    return file_path


def load_reference_target():
    reference_dict = dict()
    file_path = get_file_path_reference()
    if os.path.exists(file_path) is False:
        return reference_dict

    with open(file_path, 'r', encoding="utf-8") as f:
        reference_dict = json.load(f)

    return reference_dict


def save_reference(reference_dict):
    file_path = get_file_path_reference()
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(reference_dict, f, indent=4, sort_keys=False, ensure_ascii=False)


def check_reference(source_value, reference_dict):
    found = False
    if source_value in reference_dict:
        found = True
        return found, reference_dict[source_value]
    return found, ''


def process_translation_by_dir(dir_excel_target, cache_dict):
    file_list = sorted(os.listdir(dir_excel_target))
    for file_one in file_list:
        if file_one.endswith(FILE_EXCEL_EXT) is True:
            file_path = os.path.join(dir_excel_target, file_one)
            asyncio.run(process_translation(file_path, cache_dict))


async def process_translation(file_path_input, cache_dict):
    translator = Translator()
    wb = openpyxl.load_workbook(filename=file_path_input)

    for sheet_name in wb.sheetnames:
        sheet_name_new = f"{sheet_name}.영문"
        if sheet_name_new in wb.sheetnames:
            print(f"This sheet already translated. {file_path_input}")
            continue

        if '영문' in sheet_name:
            print(f"영문 sheet will be skipped. {file_path_input}")
            continue

        if sheet_name.endswith('영문') is True:
            print(f"영문 sheet will be skipped. {file_path_input}")
            continue

        print(f"Translated now -{sheet_name}")
        data_dict = dict()
        ws_input = wb[sheet_name]

        ws_output = wb.copy_worksheet(ws_input)
        ws_output.title = sheet_name_new

        for row in ws_input.iter_rows(min_row=ws_input.min_row, max_row=ws_input.max_row):
            for cell in row:
                row_index = cell.row
                col_letter = cell.column_letter
                if cell.value is None:
                    continue
                value_str = str(cell.value)
                if value_str is not None:
                    cell_loc = f'{col_letter}{row_index}'
                    print(f"[{cell_loc}]-Original-{value_str}")
                    reference_found, translated = check_reference(value_str, cache_dict)
                    if reference_found is False:
                        translated_result = await translator.translate(value_str, dest='en')
                        translated = translated_result.text
                        cache_dict[value_str] = translated
                        print(f"[{cell_loc}]-Translated-{translated}")
                    else:
                        print(f"[{cell_loc}]-Existed-{translated}")

                    data_dict[cell_loc] = translated

                    new_cell = ws_output.cell(row=cell.row, column=cell.col_idx, value=translated)
                    if cell.has_style:
                        font_new = copy.copy(cell.font)
                        font_new.name = FONT_NAME_EN
                        new_cell.font = font_new
                        new_cell.border = copy.copy(cell.border)
                        new_cell.fill = copy.copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy.copy(cell.protection)
                        new_cell.alignment = copy.copy(cell.alignment)

    save_reference(cache_dict)

    wb.save(file_path_input)

    return "OK"


if __name__ == '__main__':
    reference_dict = load_reference_target()
    process_translation_by_dir(DIR_EXCEL_TARGET, reference_dict)
